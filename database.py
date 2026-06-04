import sqlite3
import os
import io
import calendar
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "inventario.db")


def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


# ── Schema ──────────────────────────────────────────────────────

def init_db():
    conn = get_conn()
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS areas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT UNIQUE NOT NULL
    );

    CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        nombre TEXT NOT NULL,
        rol TEXT NOT NULL DEFAULT 'empleado',
        area_id INTEGER,
        activo INTEGER NOT NULL DEFAULT 1,
        FOREIGN KEY (area_id) REFERENCES areas(id)
    );

    CREATE TABLE IF NOT EXISTS productos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        categoria TEXT NOT NULL DEFAULT 'General',
        unidad TEXT NOT NULL DEFAULT 'PIEZA',
        activo INTEGER NOT NULL DEFAULT 1
    );

    CREATE TABLE IF NOT EXISTS ciclos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        fecha_inicio TEXT NOT NULL,
        fecha_cierre TEXT NOT NULL,
        estado TEXT NOT NULL DEFAULT 'abierto'
    );

    CREATE TABLE IF NOT EXISTS pedidos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario_id INTEGER NOT NULL,
        area_id INTEGER NOT NULL,
        ciclo_id INTEGER NOT NULL,
        fecha_pedido TEXT NOT NULL,
        estado TEXT NOT NULL DEFAULT 'pendiente',
        notas TEXT,
        FOREIGN KEY (usuario_id) REFERENCES usuarios(id),
        FOREIGN KEY (area_id) REFERENCES areas(id),
        FOREIGN KEY (ciclo_id) REFERENCES ciclos(id)
    );

    CREATE TABLE IF NOT EXISTS detalle_pedido (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        pedido_id INTEGER NOT NULL,
        producto_id INTEGER NOT NULL,
        cantidad INTEGER NOT NULL,
        FOREIGN KEY (pedido_id) REFERENCES pedidos(id),
        FOREIGN KEY (producto_id) REFERENCES productos(id)
    );
    """)

    # ── Seed areas ──
    areas = [
        "Talento Humano",
        "Mesa de Atencion al Cliente",
        "Contabilidad",
        "Finanzas",
        "Credito y Cobranza",
        "Abastecimiento y Compras",
        "Tecnologia de Informacion (TI)",
        "Business Intelligence (BI)",
    ]
    for nombre in areas:
        conn.execute("INSERT OR IGNORE INTO areas (nombre) VALUES (?)", (nombre,))

    # ── Seed admin ──
    if not conn.execute("SELECT id FROM usuarios WHERE username='admin'").fetchone():
        conn.execute(
            "INSERT INTO usuarios (username, password_hash, nombre, rol) VALUES (?,?,?,?)",
            ("admin", generate_password_hash("admin123"), "Jorge Alvarez", "admin"),
        )

    conn.commit()
    conn.close()


def cargar_catalogo():
    """Carga catalogo PROESA si esta vacio."""
    conn = get_conn()
    if conn.execute("SELECT COUNT(*) as c FROM productos").fetchone()["c"] > 0:
        conn.close()
        return
    productos = [
        ("Block de vales naranja", "Papeleria", "PIEZA"),
        ("Block de vales azules", "Papeleria", "PIEZA"),
        ("Folder tamano carta beige", "Papeleria", "PIEZA"),
        ("Folder tamano oficio azul", "Papeleria", "PIEZA"),
        ("Mica pliego", "Papeleria", "PIEZA"),
        ("Protector de hojas tamano carta", "Papeleria", "PAQUETE"),
        ("Sobre para efectivo", "Papeleria", "PIEZA"),
        ("Boligrafo tinta azul", "Escritura", "PIEZA"),
        ("Boligrafo tinta negra", "Escritura", "PIEZA"),
        ("Boligrafo tinta roja", "Escritura", "PIEZA"),
        ("Boligrafo tinta verde", "Escritura", "PIEZA"),
        ("Goma", "Escritura", "PIEZA"),
        ("Lapiz", "Escritura", "PIEZA"),
        ("Marca textos amarillo", "Escritura", "PIEZA"),
        ("Marca textos azul", "Escritura", "PIEZA"),
        ("Marca textos naranja", "Escritura", "PIEZA"),
        ("Marca textos rosa", "Escritura", "PIEZA"),
        ("Marca textos verde", "Escritura", "PIEZA"),
        ("Plumon punto fino azul", "Escritura", "PIEZA"),
        ("Plumon punto fino negro", "Escritura", "PIEZA"),
        ("Plumon punto fino rojo", "Escritura", "PIEZA"),
        ("Plumones lavables para pizarron", "Escritura", "PAQUETE"),
        ("Sacapuntas", "Escritura", "PIEZA"),
        ("Clip #2", "Sujetadores", "CAJA"),
        ("Clip tipo mariposa", "Sujetadores", "CAJA"),
        ("Grapas", "Sujetadores", "CAJA"),
        ("Diurex", "Adhesivos", "PIEZA"),
        ("Lapiz adhesivo", "Adhesivos", "PIEZA"),
        ("Pegamento Kola loka", "Adhesivos", "PIEZA"),
        ("Cutter", "Herramientas", "PIEZA"),
        ("Engrapadora", "Herramientas", "PIEZA"),
        ("Quita grapas", "Herramientas", "PIEZA"),
        ("Tijeras", "Herramientas", "PIEZA"),
        ("Pilas AA", "Pilas", "PAQUETE"),
        ("Pilas AAA", "Pilas", "PAQUETE"),
        ("Pilas Cuadrada", "Pilas", "PIEZA"),
        ("Pilas D", "Pilas", "PIEZA"),
        ("Manga plastica cubre polvo", "Proteccion", "PIEZA"),
        ("Papel aluminio rollo", "Varios", "ROLLO"),
        ("Porta gafete cordon retractil", "Varios", "PIEZA"),
    ]
    for nombre, cat, unidad in productos:
        conn.execute(
            "INSERT INTO productos (nombre, categoria, unidad) VALUES (?,?,?)",
            (nombre, cat, unidad),
        )
    conn.commit()
    conn.close()


# ── Auth ────────────────────────────────────────────────────────

def verificar_login(username, password):
    conn = get_conn()
    user = conn.execute(
        "SELECT * FROM usuarios WHERE username=? AND activo=1", (username,)
    ).fetchone()
    conn.close()
    if user and check_password_hash(user["password_hash"], password):
        return dict(user)
    return None


def get_usuario(user_id):
    conn = get_conn()
    user = conn.execute(
        """SELECT u.*, a.nombre as area_nombre
           FROM usuarios u LEFT JOIN areas a ON u.area_id=a.id
           WHERE u.id=?""",
        (user_id,),
    ).fetchone()
    conn.close()
    return dict(user) if user else None


def get_usuarios():
    conn = get_conn()
    rows = conn.execute(
        """SELECT u.*, a.nombre as area_nombre
           FROM usuarios u LEFT JOIN areas a ON u.area_id=a.id
           ORDER BY u.rol, u.nombre"""
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def crear_usuario(username, password, nombre, rol, area_id):
    conn = get_conn()
    try:
        conn.execute(
            "INSERT INTO usuarios (username, password_hash, nombre, rol, area_id) VALUES (?,?,?,?,?)",
            (username, generate_password_hash(password), nombre, rol, area_id if area_id else None),
        )
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()


def actualizar_usuario(user_id, nombre, rol, area_id, activo, password=None):
    conn = get_conn()
    if password:
        conn.execute(
            "UPDATE usuarios SET nombre=?, rol=?, area_id=?, activo=?, password_hash=? WHERE id=?",
            (nombre, rol, area_id if area_id else None, activo, generate_password_hash(password), user_id),
        )
    else:
        conn.execute(
            "UPDATE usuarios SET nombre=?, rol=?, area_id=?, activo=? WHERE id=?",
            (nombre, rol, area_id if area_id else None, activo, user_id),
        )
    conn.commit()
    conn.close()


def eliminar_usuario(user_id):
    conn = get_conn()
    conn.execute("UPDATE usuarios SET activo=0 WHERE id=?", (user_id,))
    conn.commit()
    conn.close()


# ── Areas ───────────────────────────────────────────────────────

def get_areas():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM areas ORDER BY nombre").fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ── Productos ───────────────────────────────────────────────────

def get_productos(solo_activos=True):
    conn = get_conn()
    if solo_activos:
        rows = conn.execute("SELECT * FROM productos WHERE activo=1 ORDER BY categoria, nombre").fetchall()
    else:
        rows = conn.execute("SELECT * FROM productos ORDER BY categoria, nombre").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_categorias():
    conn = get_conn()
    rows = conn.execute("SELECT DISTINCT categoria FROM productos WHERE activo=1 ORDER BY categoria").fetchall()
    conn.close()
    return [r["categoria"] for r in rows]


def agregar_producto(nombre, categoria, unidad):
    conn = get_conn()
    conn.execute("INSERT INTO productos (nombre, categoria, unidad) VALUES (?,?,?)", (nombre, categoria, unidad.upper()))
    conn.commit()
    conn.close()


def actualizar_producto(prod_id, nombre, categoria, unidad, activo):
    conn = get_conn()
    conn.execute(
        "UPDATE productos SET nombre=?, categoria=?, unidad=?, activo=? WHERE id=?",
        (nombre, categoria, unidad.upper(), activo, prod_id),
    )
    conn.commit()
    conn.close()


def eliminar_producto(prod_id):
    conn = get_conn()
    conn.execute("UPDATE productos SET activo=0 WHERE id=?", (prod_id,))
    conn.commit()
    conn.close()


# ── Ciclos ──────────────────────────────────────────────────────

def asegurar_ciclo_activo():
    conn = get_conn()
    activo = conn.execute(
        "SELECT * FROM ciclos WHERE estado='abierto' ORDER BY fecha_inicio DESC LIMIT 1"
    ).fetchone()

    if activo:
        cierre = datetime.strptime(activo["fecha_cierre"], "%Y-%m-%d")
        if datetime.now() > cierre:
            conn.execute("UPDATE ciclos SET estado='cerrado' WHERE id=?", (activo["id"],))
            conn.commit()
            activo = None

    if not activo:
        hoy = datetime.now()
        inicio = hoy.replace(day=1).strftime("%Y-%m-%d")
        ultimo_dia = calendar.monthrange(hoy.year, hoy.month)[1]
        cierre = hoy.replace(day=ultimo_dia).strftime("%Y-%m-%d")
        meses = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
            5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
            9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
        }
        nombre = f"{meses[hoy.month]} {hoy.year}"
        conn.execute(
            "INSERT INTO ciclos (nombre, fecha_inicio, fecha_cierre) VALUES (?,?,?)",
            (nombre, inicio, cierre),
        )
        conn.commit()
    conn.close()


def get_ciclo_activo():
    conn = get_conn()
    row = conn.execute(
        "SELECT * FROM ciclos WHERE estado='abierto' ORDER BY fecha_inicio DESC LIMIT 1"
    ).fetchone()
    conn.close()
    return dict(row) if row else None


# ── Pedidos ─────────────────────────────────────────────────────

def crear_pedido(usuario_id, area_id, ciclo_id, items, notas=""):
    """items = [(producto_id, cantidad), ...]"""
    conn = get_conn()
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    cur = conn.execute(
        "INSERT INTO pedidos (usuario_id, area_id, ciclo_id, fecha_pedido, estado, notas) VALUES (?,?,?,?,?,?)",
        (usuario_id, area_id, ciclo_id, now, "pendiente", notas),
    )
    pedido_id = cur.lastrowid
    for producto_id, cantidad in items:
        conn.execute(
            "INSERT INTO detalle_pedido (pedido_id, producto_id, cantidad) VALUES (?,?,?)",
            (pedido_id, producto_id, cantidad),
        )
    conn.commit()
    conn.close()
    return pedido_id


def get_pedidos_usuario(usuario_id):
    conn = get_conn()
    rows = conn.execute(
        """SELECT p.*, a.nombre as area_nombre, c.nombre as ciclo_nombre
           FROM pedidos p
           JOIN areas a ON p.area_id=a.id
           JOIN ciclos c ON p.ciclo_id=c.id
           WHERE p.usuario_id=?
           ORDER BY p.fecha_pedido DESC""",
        (usuario_id,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_pedidos_area_estado(area_id, estado):
    conn = get_conn()
    rows = conn.execute(
        """SELECT p.*, u.nombre as nombre_usuario, a.nombre as area_nombre
           FROM pedidos p
           JOIN usuarios u ON p.usuario_id=u.id
           JOIN areas a ON p.area_id=a.id
           WHERE p.area_id=? AND p.estado=?
           ORDER BY p.fecha_pedido DESC""",
        (area_id, estado),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_pedidos_por_estado(estado):
    conn = get_conn()
    rows = conn.execute(
        """SELECT p.*, u.nombre as nombre_usuario, a.nombre as area_nombre
           FROM pedidos p
           JOIN usuarios u ON p.usuario_id=u.id
           JOIN areas a ON p.area_id=a.id
           WHERE p.estado=?
           ORDER BY p.fecha_pedido DESC""",
        (estado,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_todos_pedidos():
    conn = get_conn()
    rows = conn.execute(
        """SELECT p.*, u.nombre as nombre_usuario, a.nombre as area_nombre
           FROM pedidos p
           JOIN usuarios u ON p.usuario_id=u.id
           JOIN areas a ON p.area_id=a.id
           ORDER BY p.fecha_pedido DESC"""
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_detalle_pedido(pedido_id):
    conn = get_conn()
    rows = conn.execute(
        """SELECT dp.*, p.nombre as producto_nombre, p.unidad, p.categoria
           FROM detalle_pedido dp
           JOIN productos p ON dp.producto_id=p.id
           WHERE dp.pedido_id=?""",
        (pedido_id,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def actualizar_estado_pedido(pedido_id, nuevo_estado):
    conn = get_conn()
    conn.execute("UPDATE pedidos SET estado=? WHERE id=?", (nuevo_estado, pedido_id))
    conn.commit()
    conn.close()


def borrar_pedido(pedido_id):
    conn = get_conn()
    conn.execute("DELETE FROM detalle_pedido WHERE pedido_id=?", (pedido_id,))
    conn.execute("DELETE FROM pedidos WHERE id=?", (pedido_id,))
    conn.commit()
    conn.close()


def contar_pendientes_area(area_id):
    conn = get_conn()
    n = conn.execute(
        "SELECT COUNT(*) as c FROM pedidos WHERE area_id=? AND estado='pendiente'",
        (area_id,),
    ).fetchone()["c"]
    conn.close()
    return n


def contar_pendientes_admin():
    conn = get_conn()
    n = conn.execute("SELECT COUNT(*) as c FROM pedidos WHERE estado='aprobado_lider'").fetchone()["c"]
    conn.close()
    return n


# ── Excel Export ────────────────────────────────────────────────

def generar_excel_orden(pedidos_ids, no_inventario="GA-GE-OF-00001"):
    """Genera Excel con formato de orden de compra para el proveedor."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Orden de Compra"

    # Estilos
    azul = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    fuente_header = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    fuente_data = Font(size=11, name="Calibri")
    borde = Border(
        left=Side(style="thin", color="1A5276"),
        right=Side(style="thin", color="1A5276"),
        top=Side(style="thin", color="1A5276"),
        bottom=Side(style="thin", color="1A5276"),
    )
    centro = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = ["No. Inventario", "Descripcion de Linea", "UM", "Cant. Orden"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = azul
        cell.font = fuente_header
        cell.alignment = centro
        cell.border = borde

    # Data — consolidar items iguales
    conn = get_conn()
    consolidado = {}
    for pid in pedidos_ids:
        items = conn.execute(
            """SELECT p.nombre, p.unidad, dp.cantidad
               FROM detalle_pedido dp JOIN productos p ON dp.producto_id=p.id
               WHERE dp.pedido_id=?""",
            (pid,),
        ).fetchall()
        for item in items:
            key = (item["nombre"], item["unidad"])
            consolidado[key] = consolidado.get(key, 0) + item["cantidad"]
    conn.close()

    row = 2
    for (nombre, unidad), cantidad in sorted(consolidado.items()):
        ws.cell(row=row, column=1, value=no_inventario).font = fuente_data
        ws.cell(row=row, column=2, value=nombre.upper()).font = fuente_data
        ws.cell(row=row, column=3, value=unidad.upper()).font = fuente_data
        cell_cant = ws.cell(row=row, column=4, value=float(cantidad))
        cell_cant.font = fuente_data
        cell_cant.number_format = "0.00"
        cell_cant.alignment = centro
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = borde
            ws.cell(row=row, column=c).alignment = centro if c != 2 else Alignment(vertical="center")
        row += 1

    # Anchos
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
